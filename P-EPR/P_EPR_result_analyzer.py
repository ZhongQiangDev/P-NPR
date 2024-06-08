import os
import csv
import openpyxl
import json
import xlsxwriter as xw
from tqdm import tqdm


def npr_test_results_sort(true_labels, predict_labels, topN):
    correct = 0
    total = 0
    for (t_labels, p_labels) in zip(true_labels, predict_labels):
        total += 1
        sorted_indices = sorted(range(len(p_labels)), key=lambda x: p_labels[x], reverse=True)
        flag_1 = False
        for i in range(topN):
            index = sorted_indices[i]
            if t_labels[index] == 1:
                flag_1 = True
                break
            else:
                continue
        if flag_1:
            correct += 1

    return correct, total


def npr_test_results_sort_repaired(true_labels, predict_labels, topN):
    correct = 0
    total = 0
    for (t_labels, p_labels) in zip(true_labels, predict_labels):
        if t_labels[5] == 0:
            total += 1
            sorted_indices = sorted(range(len(p_labels)), key=lambda x: p_labels[x], reverse=True)
            flag_1 = False
            for i in range(topN):
                index = sorted_indices[i]
                if t_labels[index] == 1:
                    flag_1 = True
                    break
                else:
                    continue
            if flag_1:
                correct += 1

    return correct, total


def save_results_xlsx(xlsx_path, bugids, predict_labels):
    workbook = xw.Workbook(xlsx_path)

    worksheet6 = workbook.add_worksheet("top6")
    worksheet6.activate()
    title = ['bugid', 'pred_1', 'pred_2', 'pred_3', 'pred_4', 'pred_5', 'pred_6']
    worksheet6.write_row('A1', title)
    i = 2
    for bugid, p_labels in zip(bugids, predict_labels):
        sorted_indices = sorted(range(len(p_labels)), key=lambda x: p_labels[x], reverse=True)
        index_list = sorted_indices[0:6]
        row = 'A' + str(i)
        predict_nprs = [bugid]

        for index in index_list:
            if index == 0:
                predict_nprs.append('recoder')
            elif index == 1:
                predict_nprs.append('tare')
            elif index == 2:
                predict_nprs.append('rewardrepair')
            elif index == 3:
                predict_nprs.append('selfapr')
            elif index == 4:
                predict_nprs.append('gamma')
            elif index == 5:
                predict_nprs.append('allfailure')

            #  predict_nprs.append(p_labels[index])
        worksheet6.write_row(row, predict_nprs)
        i += 1

    workbook.close()
    print(xlsx_path + ' is saved successfully!')


valid_truth = {}
workbook = openpyxl.load_workbook('P-EPR/test5000_meta.xlsx')
worksheet = workbook.active
index = 0
for row in worksheet.iter_rows():
    if index == 0:
        index += 1
        continue
    row_data = [cell.value for cell in row]
    bugid = row_data[0]
    valid_truth[bugid] = [int(row_data[1]), int(row_data[2]), int(row_data[3]), int(row_data[4]), int(row_data[5]), int(row_data[6])]

test_truth = {}
with open('P-EPR/test_code_rename.csv', 'r', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter='\t')
    i = 0
    for row in reader:
        if i == 0:
            i += 1
            continue
        test_truth[row[0]] = [int(row[2]), int(row[3]), int(row[4]), int(row[5]), int(row[6]), int(row[7])]

valid_folder = './P-EPR/results-test5000'
valid_json_list = os.listdir(valid_folder)
test_folder = './P-EPR/results-defects4j'
test_json_list = os.listdir(test_folder)

valid_predict = {}
for valid_json in tqdm(valid_json_list):
    bugid = valid_json.split('.')[0].split('_')[1]
    with open(os.path.join(valid_folder, valid_json), 'r', encoding='UTF-8') as jsonfile:
        data = json.load(jsonfile)
        valid_predict[bugid] = [data['recoder'], data['tare'], data['rewardrepair'], data['selfapr'], data['gamma'], data['allfailure']]

test_predict = {}
for test_json in tqdm(test_json_list):
    bugid = test_json.split('.')[0].split('_')[1]
    with open(os.path.join(test_folder, test_json), 'r', encoding='UTF-8') as jsonfile:
        data = json.load(jsonfile)
        test_predict[bugid] = [data['recoder'], data['tare'], data['rewardrepair'], data['selfapr'], data['gamma'], data['allfailure']]

valid_bugids = []
valid_true_labels = []
valid_predict_labels = []
for key in valid_truth.keys():
    valid_bugids.append(key)
    valid_true_labels.append(valid_truth[key])
    valid_predict_labels.append(valid_predict[key])
with open('P-EPR/valid_output.txt', 'w', encoding='UTF-8') as f:
    for topN in range(1, 7):
        if topN == 6:
            valid_correct, valid_total = npr_test_results_sort(valid_true_labels, valid_predict_labels, topN)
            f.write("Top{} - valid acc: {}/{}.".format(topN, valid_correct, valid_total) + '\n')
        else:
            valid_correct, valid_total = npr_test_results_sort(valid_true_labels, valid_predict_labels, topN)
            valid_correct_repair, valid_total_repair = npr_test_results_sort_repaired(valid_true_labels, valid_predict_labels, topN)
            f.write("Top{} - valid acc: {}/{}, {}/{}.".format(topN, valid_correct, valid_total, valid_correct_repair, valid_total_repair) + '\n')
save_results_xlsx('P-EPR/valid_results_value.xlsx', valid_bugids, valid_predict_labels)

test_bugids = []
test_true_labels = []
test_predict_labels = []
for key in test_truth.keys():
    test_bugids.append(key)
    test_true_labels.append(test_truth[key])
    test_predict_labels.append(test_predict[key])
with open('P-EPR/test_output.txt', 'w', encoding='UTF-8') as f:
    for topN in range(1, 7):
        if topN == 6:
            test_correct, test_total = npr_test_results_sort(test_true_labels, test_predict_labels, topN)
            f.write("Top{} - test acc: {}/{}.".format(topN, test_correct, test_total) + '\n')
        else:
            test_correct, test_total = npr_test_results_sort(test_true_labels, test_predict_labels, topN)
            test_correct_repair, test_total_repair = npr_test_results_sort_repaired(test_true_labels, test_predict_labels, topN)
            f.write("Top{} - test acc: {}/{}, {}/{}".format(topN, test_correct, test_total, test_correct_repair, test_total_repair) + '\n')
save_results_xlsx('P-EPR/test_results_value.xlsx', test_bugids, test_predict_labels)
